from pywhatkit import sendwhatmsg_instantly
from openpyxl import load_workbook

msg = """Hey there ,

We would like to remind you about your Interview on Tuesday 11 Oct

On the day of the Interview please head to the booth in front of engineering building B between 4:30 and 6:30 PM where you can find our organisers and register your name with them by showing them this confirmation mail with the below details as soon as you reach the cafeteria.

If you have any problems don't hesitate to contact us
at roben.club/contact

HR Team 
RobEn"""

wb = load_workbook('Recruit-2022-10-10.xlsx')
sh1 = wb['Tablib Dataset']
for i in range(2,67):
    
    Numbr = '+20'+str(sh1.cell(i,6).value)
    sendwhatmsg_instantly(Numbr,msg,10,True,5)
